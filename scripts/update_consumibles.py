#!/usr/bin/env python3
"""
update_consumibles.py
---------------------
Adds a "CONSUMIBLES DETALLE" sheet (as 2nd sheet) to Plan_Operaciones_M4.xlsx,
then updates PREMISAS Table 5 so each specialty total references this new sheet.
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

WORKBOOK_PATH = "/home/user/CLAUDE-CODE-GRAMA/Plan_Operaciones_M4.xlsx"

# ── Brand colours ──────────────────────────────────────────────────────────────
DARK   = "FF043941"   # section headers
GREEN  = "FF02D47E"   # TOTAL rows
TEAL   = "FF0A7F8F"   # summary header
LIGHT  = "FFE3F8FB"   # alternating light row
WHITE  = "FFFFFFFF"
LGRAY  = "FFF2F2F2"   # alternating dark-ish row

# ── Consumables data ──────────────────────────────────────────────────────────
# Each entry: (equipo, consumible, unidad, cantidad, precio_unit)
SPECIALTIES = [
    {
        "name": "Cocina y Repostería",
        "items": [
            ("Horno convector",       "Gas GLP (uso operativo)",        "balón 10kg",    0.5,  42.00),
            ("Batidora industrial",    "Harina sin preparar",            "kg",            8,    3.20),
            ("Batidora industrial",    "Azúcar blanca",                  "kg",            4,    4.50),
            ("Batidora industrial",    "Mantequilla",                    "kg",            2,   18.00),
            ("Batidora industrial",    "Huevos",                         "docena",        3,    7.00),
            ("Cocina industrial",      "Aceite vegetal",                 "litro",         2,    8.00),
            ("Cocina industrial",      "Sal fina",                       "kg",            1,    2.00),
            ("Cocina industrial",      "Especias básicas (kit)",         "kit",           1,   25.00),
            ("Selladora/envasadora",   "Bolsas de sellado 20x30cm",      "paquete x100",  1,   28.00),
            ("Selladora/envasadora",   "Etiquetas adhesivas",            "rollo x500",    1,   15.00),
            ("Mesas de trabajo",       "Guantes descartables",           "caja x100",     0.5, 18.00),
            ("Mesas de trabajo",       "Gorros descartables",            "paquete x100",  0.5, 12.00),
            ("Mesas de trabajo",       "Mascarillas",                    "paquete x50",   1,   12.00),
        ],
    },
    {
        "name": "Computación e Informática",
        "items": [
            ("PCs / Laptops",              "Pasta térmica",                       "tubo 100g",    3,  18.00),
            ("PCs / Laptops",              "Alcohol isopropílico 99%",            "frasco 250ml", 2,  12.00),
            ("Switch / Router",            "Cable UTP Cat6",                      "metro",        20,  1.50),
            ("Switch / Router",            "Conectores RJ45 Cat6",                "bolsa x100",   1,  35.00),
            ("Switch / Router",            "Ponchadora RJ45 (reemplazo cuchilla)","pieza",        1,  15.00),
            ("Impresora",                  "Papel bond A4",                       "resma 500h",   1,  22.00),
            ("Herramientas ensamblaje",    "Pulsera antiestática",                "unidad",       5,  12.00),
            ("Herramientas ensamblaje",    "Bridas surtidas",                     "paquete x100", 1,   8.00),
        ],
    },
    {
        "name": "Construcciones Metálicas",
        "items": [
            ("Soldadora MIG/MAG",        "Alambre MIG 0.8mm",           "rollo 5kg",    1,  120.00),
            ("Soldadora MIG/MAG",        "Gas CO2 (mezcla)",            "balón 10kg",   1,   85.00),
            ("Soldadora eléctrica (SMAW)","Electrodos 6011 3.2mm",      "kg",           4,   12.00),
            ("Soldadora eléctrica (SMAW)","Electrodos 7018 3.2mm",      "kg",           4,   14.00),
            ("Soldadora TIG",            "Varillas TIG ER70S 2.4mm",   "kg",            1,   45.00),
            ("Soldadora TIG",            "Gas argón",                   "balón 10m³",   1,  120.00),
            ("Cortadora plasma",         "Electrodos de plasma",        "paquete x5",   2,   35.00),
            ("Cortadora plasma",         "Toberas de plasma",           "paquete x5",   2,   40.00),
            ("Amoladora angular",        "Disco de corte 4.5\"",        "paquete x10",  2,   35.00),
            ("Amoladora angular",        "Disco de desbaste 4.5\"",     "paquete x5",   2,   28.00),
            ("Taladro de banco",         "Brocas HSS surtido",          "set",          1,   45.00),
            ("Material de práctica",     "Platina/ángulo acero",        "kg",          20,    4.50),
        ],
    },
    {
        "name": "Ebanistería",
        "items": [
            ("Sierra circular de mesa", "Hoja sierra 60T 10\"",         "unidad",        1,  65.00),
            ("Sierra circular de mesa", "MDF 15mm 1.83x2.44m",         "plancha",       3,  78.00),
            ("Sierra de cinta",         "Hoja de banda (reemplazo)",    "unidad",        1,  68.00),
            ("Router / CNC",            "Fresa recta 6mm carburo",      "unidad",        3,  38.00),
            ("Router / CNC",            "Fresa de perfil 1/4\"",        "unidad",        2,  42.00),
            ("Lijadora orbital",        "Disco lija 80 5\"",            "paquete x20",   1,  22.00),
            ("Lijadora orbital",        "Disco lija 120 5\"",           "paquete x20",   1,  22.00),
            ("Lijadora orbital",        "Disco lija 220 5\"",           "paquete x20",   1,  25.00),
            ("Torno de madera",         "Lija para torno (rollo)",      "metro",         5,   4.00),
            ("Compresor + pistola",     "Sellador madera",              "litro",         1,  28.00),
            ("Compresor + pistola",     "Barniz brillante",             "litro",         1,  32.00),
            ("Compresor + pistola",     "Thinner",                      "litro",         1,  12.00),
            ("Material de práctica",    "Tablas tornillo 1\"x4\"x8'",  "unidad",        8,  18.00),
        ],
    },
    {
        "name": "Electricidad",
        "items": [
            ("Tablero de prácticas",    "Cable THW 14AWG (rojo)",           "metro",        20,  1.80),
            ("Tablero de prácticas",    "Cable THW 14AWG (negro)",          "metro",        20,  1.80),
            ("Tablero de prácticas",    "Cable THW 14AWG (verde)",          "metro",        10,  1.80),
            ("Tablero de prácticas",    "Terminales preaislados surtido",   "bolsa x100",   1,  28.00),
            ("Módulo automatización",   "Relé 24VDC con base",              "unidad",       5,  12.00),
            ("Módulo automatización",   "Contactor LC1-D09",                "unidad",       3,  42.00),
            ("Módulo automatización",   "Cinta aislante 19mm",              "rollo",        5,   4.00),
            ("Instalaciones práctica",  "Tubería PVC 3/4\"",               "metro",        10,  3.20),
            ("Instalaciones práctica",  "Curvas PVC 3/4\"",                "unidad",       10,  1.50),
            ("Instalaciones práctica",  "Cajas de pase 100x100",           "unidad",        5,  4.50),
            ("Equipo de medición",      "Fusibles surtido",                 "paquete x20",  2,  12.00),
        ],
    },
    {
        "name": "Electrónica",
        "items": [
            ("Estación de soldadura",   "Estaño 60/40 1mm",             "rollo 250g",   1,  32.00),
            ("Estación de soldadura",   "Flux líquido",                 "frasco 50ml",  2,  18.00),
            ("Estación de soldadura",   "Esponja de limpieza",          "unidad",       3,   5.00),
            ("Estación de aire (SMD)",  "Pasta soldadora SMD",          "jeringa 50g",  1,  38.00),
            ("Estación de aire (SMD)",  "Mecha desoldar 2.5mm",         "rollo x3",     1,  20.00),
            ("Protoboard / PCB",        "PCB virgen fibra 10x15cm",     "unidad",      10,   5.50),
            ("Protoboard / PCB",        "Resistencias surtido",         "kit 600u",     1,  22.00),
            ("Protoboard / PCB",        "Condensadores surtido",        "kit 200u",     1,  28.00),
            ("Protoboard / PCB",        "Transistores BC547/BC557",     "paquete x50",  1,  18.00),
            ("Protoboard / PCB",        "ICs básicos (555, LM358, 7805)","kit",         1,  35.00),
            ("Protoboard / PCB",        "Cables jumper M-M 20cm",       "paquete x40",  2,  12.00),
            ("Fuente de alimentación",  "Cables banana-cocodrilo",      "par",          4,   8.00),
        ],
    },
    {
        "name": "Industria Alimentaria",
        "items": [
            ("Mezcladora industrial",   "Harina preparada",             "kg",            6,  4.50),
            ("Mezcladora industrial",   "Azúcar blanca",                "kg",            4,  4.50),
            ("Mezcladora industrial",   "Sal",                          "kg",            1,  2.00),
            ("Pasteurizador",           "Leche fresca",                 "litro",        10,  4.50),
            ("Pasteurizador",           "Envases de vidrio 250ml",      "unidad",       20,  3.50),
            ("Selladora al vacío",      "Bolsas vacío 20x30cm",         "paquete x100",  1, 35.00),
            ("Etiquetadora",            "Etiquetas en blanco",          "rollo x500",    1, 18.00),
            ("Mesas acero inox",        "Guantes descartables",         "caja x100",     1, 18.00),
            ("Mesas acero inox",        "Gorros descartables",          "paquete x100",  1, 12.00),
            ("Mesas acero inox",        "Mascarillas",                  "paquete x50",   1, 12.00),
            ("Mesas acero inox",        "Detergente industrial",        "kg",            1, 15.00),
        ],
    },
    {
        "name": "Industria del Vestido",
        "items": [
            ("Máq. costura industrial", "Hilo cono Nro.40 (blanco)",    "cono",          2, 12.00),
            ("Máq. costura industrial", "Hilo cono Nro.40 (colores)",   "cono",          3, 12.00),
            ("Máq. costura industrial", "Agujas industriales 14/90",    "caja x10",      2, 18.00),
            ("Overlock (remalladora)",  "Hilos overlock x4 colores",    "cono",          8, 10.00),
            ("Overlock (remalladora)",  "Agujas overlock",              "paquete x10",   1, 14.00),
            ("Recubridora",             "Hilos recubridora",            "cono",          6, 10.00),
            ("Recubridora",             "Agujas recubridora",           "paquete x5",    1, 16.00),
            ("Cortadora circular",      "Hoja de corte (reemplazo)",    "unidad",        2, 45.00),
            ("Plancha industrial",      "Agua destilada",               "litro",         2,  4.00),
            ("Material de práctica",    "Tela popelina básica",         "metro",         8,  8.00),
            ("Material de práctica",    "Tela drill 14oz",              "metro",         4, 12.00),
        ],
    },
    {
        "name": "Mecánica Automotriz",
        "items": [
            ("Elevador hidráulico",         "Aceite hidráulico ISO 46",     "litro",         1,  22.00),
            ("Desmontadora neumáticos",     "Lubricante de montaje",        "litro",         1,  18.00),
            ("Balanceadora",                "Pesas adhesivas balanceo",     "caja 100u",     1,  48.00),
            ("Compresor de aire",           "Aceite para compresor",        "litro",        0.5, 25.00),
            ("Práctica cambio aceite",      "Aceite motor 15W40",           "litro",         4,  14.00),
            ("Práctica cambio aceite",      "Filtro de aceite",             "unidad",        2,  28.00),
            ("Práctica cambio aceite",      "Trapo industrial",             "kg",            1,  12.00),
            ("Práctica de frenos",          "Pastillas de freno (par)",     "par",           1,  48.00),
            ("Práctica de frenos",          "Líquido de frenos DOT4",       "frasco 500ml",  1,  22.00),
            ("Práctica de frenos",          "Limpiador de frenos",          "spray 400ml",   2,  18.00),
            ("Escáner diagnóstico",         "Cable OBD2 repuesto",          "unidad",        1,  35.00),
        ],
    },
]

# ── Style helpers ──────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="FF000000", size=11):
    return Font(bold=bold, color=color, size=size)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    s = Side(style="thin", color="FFD0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

MONEY_FMT = '#,##0.00'
NUM_FMT   = '0.00'

# ── Build the CONSUMIBLES DETALLE sheet ───────────────────────────────────────
def build_consumibles_sheet(wb):
    ws = wb.create_sheet("CONSUMIBLES DETALLE")

    # No gridlines
    ws.sheet_view.showGridLines = False

    # Column widths (A=28, B=30, C=16, D=10, E=16, F=16)
    col_widths = {"A": 28, "B": 30, "C": 16, "D": 10, "E": 16, "F": 16}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Page title
    ws.row_dimensions[1].height = 28
    ws["A1"] = "CONSUMIBLES DETALLE — COSTO UNITARIO POR EQUIPO Y ESPECIALIDAD"
    ws["A1"].font = Font(bold=True, color="FF" + DARK[2:], size=13)
    ws["A1"].alignment = align("left", "center")

    current_row = 2
    specialty_total_cells = {}   # name -> cell address of TOTAL in col F

    for spec in SPECIALTIES:
        spec_name = spec["name"]
        items     = spec["items"]

        # ── Section header ─────────────────────────────────────────────────
        ws.row_dimensions[current_row].height = 22
        header_cells = [ws.cell(row=current_row, column=c) for c in range(1, 7)]
        for cell in header_cells:
            cell.fill      = fill(DARK)
            cell.font      = font(bold=True, color=WHITE, size=11)
            cell.alignment = align("left", "center")
        ws.cell(row=current_row, column=1).value = spec_name.upper()
        current_row += 1

        # ── Column sub-headers ─────────────────────────────────────────────
        subheaders = ["Equipo", "Consumible", "Unidad",
                      "Cantidad", "Precio Unit. (S/.)", "Subtotal"]
        for c, label in enumerate(subheaders, 1):
            cell = ws.cell(row=current_row, column=c)
            cell.value     = label
            cell.fill      = fill(TEAL)
            cell.font      = font(bold=True, color=WHITE, size=10)
            cell.alignment = align("center", "center")
            cell.border    = thin_border()
        current_row += 1

        # ── Data rows ─────────────────────────────────────────────────────
        first_data_row = current_row
        for i, (equipo, consumible, unidad, qty, price) in enumerate(items):
            row_fill = fill(LGRAY) if i % 2 == 0 else fill(WHITE)
            for c in range(1, 7):
                cell = ws.cell(row=current_row, column=c)
                cell.fill      = row_fill
                cell.alignment = align("left", "center")
                cell.border    = thin_border()
            ws.cell(row=current_row, column=1).value = equipo
            ws.cell(row=current_row, column=2).value = consumible
            ws.cell(row=current_row, column=3).value = unidad
            ws.cell(row=current_row, column=3).alignment = align("center", "center")

            qty_cell   = ws.cell(row=current_row, column=4)
            qty_cell.value  = qty
            qty_cell.number_format = NUM_FMT
            qty_cell.alignment = align("center", "center")

            price_cell = ws.cell(row=current_row, column=5)
            price_cell.value  = price
            price_cell.number_format = MONEY_FMT
            price_cell.alignment = align("right", "center")

            sub_cell   = ws.cell(row=current_row, column=6)
            sub_cell.value  = f"=D{current_row}*E{current_row}"
            sub_cell.number_format = MONEY_FMT
            sub_cell.alignment = align("right", "center")

            current_row += 1

        last_data_row = current_row - 1

        # ── TOTAL row ─────────────────────────────────────────────────────
        ws.row_dimensions[current_row].height = 20
        for c in range(1, 7):
            cell = ws.cell(row=current_row, column=c)
            cell.fill      = fill(GREEN)
            cell.font      = font(bold=True, color="FF043941", size=11)
            cell.alignment = align("center", "center")
            cell.border    = thin_border()
        ws.cell(row=current_row, column=1).value = "TOTAL"
        ws.cell(row=current_row, column=1).alignment = align("left", "center")
        ws.cell(row=current_row, column=2).value = spec_name
        ws.cell(row=current_row, column=2).alignment = align("left", "center")

        total_cell = ws.cell(row=current_row, column=6)
        total_cell.value  = f"=SUM(F{first_data_row}:F{last_data_row})"
        total_cell.number_format = MONEY_FMT

        specialty_total_cells[spec_name] = f"F{current_row}"

        current_row += 2   # blank row between specialties

    # ── Summary table ─────────────────────────────────────────────────────────
    current_row += 1   # extra blank before summary
    summary_start = current_row

    # Summary header
    ws.row_dimensions[current_row].height = 22
    for c, label in enumerate(["Especialidad", "Total Consumibles (S/.)"], 1):
        cell = ws.cell(row=current_row, column=c)
        cell.value     = label
        cell.fill      = fill(TEAL)
        cell.font      = font(bold=True, color=WHITE, size=11)
        cell.alignment = align("center", "center")
        cell.border    = thin_border()
    current_row += 1

    grand_total_refs = []
    for i, spec in enumerate(SPECIALTIES):
        spec_name = spec["name"]
        row_fill  = fill(LGRAY) if i % 2 == 0 else fill(WHITE)
        for c in range(1, 3):
            cell = ws.cell(row=current_row, column=c)
            cell.fill      = row_fill
            cell.border    = thin_border()
            cell.alignment = align("center", "center")
        ws.cell(row=current_row, column=1).value     = spec_name
        ws.cell(row=current_row, column=1).alignment = align("left", "center")
        ref_cell = ws.cell(row=current_row, column=2)
        ref_cell.value         = f"={specialty_total_cells[spec_name]}"
        ref_cell.number_format = MONEY_FMT

        grand_total_refs.append(f"B{current_row}")
        current_row += 1

    # Grand total row
    ws.row_dimensions[current_row].height = 22
    for c in range(1, 3):
        cell = ws.cell(row=current_row, column=c)
        cell.fill      = fill(DARK)
        cell.font      = font(bold=True, color=WHITE, size=11)
        cell.alignment = align("center", "center")
        cell.border    = thin_border()
    ws.cell(row=current_row, column=1).value     = "GRAN TOTAL CONSUMIBLES"
    ws.cell(row=current_row, column=1).alignment = align("left", "center")
    grand_total_cell = ws.cell(row=current_row, column=2)
    grand_total_cell.value         = f"=SUM({grand_total_refs[0]}:{grand_total_refs[-1]})"
    grand_total_cell.number_format = MONEY_FMT

    return ws, specialty_total_cells


# ── Update PREMISAS Table 5 ───────────────────────────────────────────────────
def update_premisas_table5(wb, specialty_total_cells):
    ws = wb["PREMISAS"]
    # Table 5 rows: B44:C52 (row 44 = Cocina y Repostería … row 52 = Mecánica Automotriz)
    # Order matches SPECIALTIES list exactly.
    start_row = 44
    for i, spec in enumerate(SPECIALTIES):
        row    = start_row + i
        spec_name = spec["name"]
        tcell  = specialty_total_cells[spec_name]   # e.g. "F15"
        ws.cell(row=row, column=3).value = f"='CONSUMIBLES DETALLE'!{tcell}"


# ── Move sheet to position 2 (index 1) ───────────────────────────────────────
def move_sheet_to_second(wb, sheet_name):
    sheet_names = wb.sheetnames
    if sheet_name not in sheet_names:
        raise ValueError(f"Sheet '{sheet_name}' not found")
    # openpyxl low-level: move sheet
    idx = sheet_names.index(sheet_name)
    if idx == 1:
        return   # already in place
    wb.move_sheet(sheet_name, offset=1 - idx)


# ── Compute grand total manually (for printing) ───────────────────────────────
def compute_grand_total():
    total = 0.0
    for spec in SPECIALTIES:
        for _eq, _cons, _unit, qty, price in spec["items"]:
            total += qty * price
    return total


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    wb = openpyxl.load_workbook(WORKBOOK_PATH)

    # Remove existing sheet if present (idempotent)
    if "CONSUMIBLES DETALLE" in wb.sheetnames:
        del wb["CONSUMIBLES DETALLE"]

    _ws, specialty_total_cells = build_consumibles_sheet(wb)

    # Move to position 2 (after PREMISAS, before DETALLE M4)
    move_sheet_to_second(wb, "CONSUMIBLES DETALLE")

    # Update PREMISAS Table 5
    update_premisas_table5(wb, specialty_total_cells)

    wb.save(WORKBOOK_PATH)

    grand_total = compute_grand_total()
    print("=" * 55)
    print("CONSUMIBLES DETALLE sheet created successfully.")
    print(f"Sheet order: {wb.sheetnames}")
    print("-" * 55)
    for spec in SPECIALTIES:
        sub = sum(qty * price for _, _, _, qty, price in spec["items"])
        print(f"  {spec['name']:<30}  S/. {sub:>8,.2f}")
    print("-" * 55)
    print(f"  {'GRAN TOTAL':<30}  S/. {grand_total:>8,.2f}")
    print("=" * 55)
    print("PREMISAS Table 5 updated with cell references.")


if __name__ == "__main__":
    main()
