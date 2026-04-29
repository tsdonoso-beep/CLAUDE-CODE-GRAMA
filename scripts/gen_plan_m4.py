#!/usr/bin/env python3
"""
gen_plan_m4.py
Generates Plan_Operaciones_M4.xlsx with 4 sheets:
  PREMISAS, DETALLE M4, RESUMEN IE, RESUMEN ESPECIALIDAD
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.table import Table, TableStyleInfo
import os

# ---------------------------------------------------------------------------
# BRAND COLORS
# ---------------------------------------------------------------------------
C_DARK   = "043941"
C_GREEN  = "02D47E"
C_TEAL   = "0A7F8F"
C_LIGHT  = "E3F8FB"
C_YELLOW = "FFFF99"
C_WHITE  = "FFFFFF"
C_ALT    = "EAF7F9"   # very light teal for alternating rows
C_HEADER = "D0EFF4"   # column header row

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color="000000", size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border():
    side = Side(style="thin", color="CCCCCC")
    return Border(left=side, right=side, top=side, bottom=side)

def bottom_border():
    side = Side(style="medium", color=C_TEAL)
    return Border(bottom=side)

# ---------------------------------------------------------------------------
# DATA
# ---------------------------------------------------------------------------
TALLERES = [
    (1,  "T-COCINA-01",   "16449 Eloy Soberón Flores",  "IE1", "Cajamarca – San Ignacio",        "C", "Cocina y Repostería",       "NO"),
    (2,  "T-COMP-02",     "2026 Simón Bolívar",          "IE2", "Lima Metropolitana – Comas",     "A", "Computación e Informática", "NO"),
    (3,  "T-ELEC-02",     "2026 Simón Bolívar",          "IE2", "Lima Metropolitana – Comas",     "A", "Electricidad",              "SI"),
    (4,  "T-INDUS.A-02",  "2026 Simón Bolívar",          "IE2", "Lima Metropolitana – Comas",     "A", "Industria Alimentaria",     "NO"),
    (5,  "T-INDUS.V-02",  "2026 Simón Bolívar",          "IE2", "Lima Metropolitana – Comas",     "A", "Industria del Vestido",     "NO"),
    (6,  "T-COMP-03",     "6049 Ricardo Palma",          "IE3", "Lima Metropolitana – Surquillo", "A", "Computación e Informática", "NO"),
    (7,  "T-ELECTRO-03",  "6049 Ricardo Palma",          "IE3", "Lima Metropolitana – Surquillo", "A", "Electrónica",               "SI"),
    (8,  "T-ELEC-03",     "6049 Ricardo Palma",          "IE3", "Lima Metropolitana – Surquillo", "A", "Electricidad",              "SI"),
    (9,  "T-INDUS.A-04",  "Francisco Irazola",           "IE4", "Junín – Satipo",                 "C", "Industria Alimentaria",     "NO"),
    (10, "T-MECA.A-04",   "Francisco Irazola",           "IE4", "Junín – Satipo",                 "C", "Mecánica Automotriz",       "SI"),
    (11, "T-COMP-05",     "Guillermo E. Billinghurst",   "IE5", "Lima Provincia – Barranca",      "B", "Computación e Informática", "NO"),
    (12, "T-EBAN-05",     "Guillermo E. Billinghurst",   "IE5", "Lima Provincia – Barranca",      "B", "Ebanistería",               "SI"),
    (13, "T-ELEC-05",     "Guillermo E. Billinghurst",   "IE5", "Lima Provincia – Barranca",      "B", "Electricidad",              "SI"),
    (14, "T-INDUS.A-05",  "Guillermo E. Billinghurst",   "IE5", "Lima Provincia – Barranca",      "B", "Industria Alimentaria",     "NO"),
    (15, "T-MECA.A-05",   "Guillermo E. Billinghurst",   "IE5", "Lima Provincia – Barranca",      "B", "Mecánica Automotriz",       "SI"),
    (16, "T-MECA.A-06",   "José Granda",                 "IE6", "Lima Metropolitana – SMP",       "A", "Mecánica Automotriz",       "SI"),
    (17, "T-COMP-06",     "José Granda",                 "IE6", "Lima Metropolitana – SMP",       "A", "Computación e Informática", "NO"),
    (18, "T-INDUS.A-06",  "José Granda",                 "IE6", "Lima Metropolitana – SMP",       "A", "Industria Alimentaria",     "NO"),
    (19, "T-INDUS.V-07",  "Manuel A. Mesones Muro",      "IE7", "San Martín – El Dorado",         "C", "Industria del Vestido",     "NO"),
    (20, "T-CONST.M-07",  "Manuel A. Mesones Muro",      "IE7", "San Martín – El Dorado",         "C", "Construcciones Metálicas",  "SI"),
    (21, "T-EBAN-07",     "Manuel A. Mesones Muro",      "IE7", "San Martín – El Dorado",         "C", "Ebanistería",               "SI"),
]

TARIFAS = [
    ("Cocina y Repostería",       25, 18),
    ("Computación e Informática", 22, 15),
    ("Construcciones Metálicas",  32, 23),
    ("Ebanistería",               28, 20),
    ("Electricidad",              28, 20),
    ("Electrónica",               32, 23),
    ("Industria Alimentaria",     22, 15),
    ("Industria del Vestido",     22, 15),
    ("Mecánica Automotriz",       32, 23),
]

GRUPOS = [
    ("A", "Lima Metropolitana",  60, 40,   0),
    ("B", "Lima Provincia",     120, 20, 120),
    ("C", "Región lejana",      160, 25, 100),
]

COLEGIOS = [
    ("IE1", "16449 Eloy Soberón Flores", "Lima–Jaén",         "Vuelo",     900),
    ("IE2", "2026 Simón Bolívar",        "Lima–Comas",        "Propio",      0),
    ("IE3", "6049 Ricardo Palma",        "Lima–Surquillo",    "Propio",      0),
    ("IE4", "Francisco Irazola",         "Lima–Satipo",       "Bus/Van",   180),
    ("IE5", "Guillermo E. Billinghurst", "Lima–Barranca",     "Bus",        60),
    ("IE6", "José Granda",               "Lima–SMP",          "Propio",      0),
    ("IE7", "Manuel A. Mesones Muro",    "Lima–Tarapoto+Bus", "Vuelo+Bus", 750),
]

CONSUMIBLES = [
    ("Cocina y Repostería",       800),
    ("Computación e Informática", 200),
    ("Construcciones Metálicas",  900),
    ("Ebanistería",               700),
    ("Electricidad",              500),
    ("Electrónica",               600),
    ("Industria Alimentaria",     500),
    ("Industria del Vestido",     350),
    ("Mecánica Automotriz",       800),
]

# ---------------------------------------------------------------------------
# HELPER: apply header style
# ---------------------------------------------------------------------------
def section_header(ws, row, col, text, span=None):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = fill(C_TEAL)
    cell.font = font(bold=True, color=C_WHITE, size=10)
    cell.alignment = align("center")
    if span:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + span - 1)
    return cell

def col_header(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.fill = fill(C_HEADER)
    cell.font = font(bold=True, size=9)
    cell.alignment = align("center", wrap=True)
    cell.border = thin_border()
    return cell

def editable_cell(ws, row, col, value, fmt=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = fill(C_YELLOW)
    cell.font = font(size=10)
    cell.alignment = align("right")
    cell.border = thin_border()
    if fmt:
        cell.number_format = fmt
    return cell

def label_cell(ws, row, col, text, bold=False):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = font(bold=bold, size=10)
    cell.alignment = align("left")
    cell.border = thin_border()
    return cell

def formula_cell(ws, row, col, formula, fmt=None):
    cell = ws.cell(row=row, column=col, value=formula)
    cell.fill = fill(C_LIGHT)
    cell.font = font(italic=True, size=10)
    cell.alignment = align("right")
    cell.border = thin_border()
    if fmt:
        cell.number_format = fmt
    return cell

def no_gridlines(ws):
    ws.sheet_view.showGridLines = False

# ---------------------------------------------------------------------------
# SHEET 1 — PREMISAS
# ---------------------------------------------------------------------------
def build_premisas(wb):
    ws = wb.create_sheet("PREMISAS")
    no_gridlines(ws)
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 22

    # Title
    ws.merge_cells("B1:F1")
    tc = ws["B1"]
    tc.value = "PREMISAS DEL PLAN DE OPERACIONES — M4"
    tc.fill = fill(C_DARK)
    tc.font = font(bold=True, color=C_WHITE, size=13)
    tc.alignment = align("center")
    ws.row_dimensions[1].height = 28

    row = 3
    # ---- TABLE 1: PARÁMETROS DE TIEMPO ----
    section_header(ws, row, 2, "TABLA 1 — PARÁMETROS DE TIEMPO", span=3)
    row += 1

    params = [
        ("dias_cap",    "Días capacitación presencial",        8,  None),
        ("horas_dia",   "Horas por día",                       5,  None),
        ("dias_prep",   "Días preparación técnico lead",       2,  None),
        ("dias_viaje",  "Días viaje ida+vuelta",                2,  None),
        ("hh_dia",      "HH/día facturables",                  8,  None),
        ("noches",      "Noches hospedaje",                    9,  None),
    ]

    # column headers
    col_header(ws, row, 2, "Parámetro")
    col_header(ws, row, 3, "Valor")
    row += 1

    param_rows = {}
    for name, label, val, fmt in params:
        label_cell(ws, row, 2, label)
        editable_cell(ws, row, 3, val, fmt)
        param_rows[name] = row
        row += 1

    # Derived formulas
    dias_cap_ref   = f"C{param_rows['dias_cap']}"
    dias_prep_ref  = f"C{param_rows['dias_prep']}"
    dias_viaje_ref = f"C{param_rows['dias_viaje']}"

    label_cell(ws, row, 2, "Total días facturables LEAD (cap+prep+viaje)")
    f_lead = formula_cell(ws, row, 3, f"={dias_cap_ref}+{dias_prep_ref}+{dias_viaje_ref}")
    param_rows["total_lead"] = row
    row += 1

    label_cell(ws, row, 2, "Total días facturables APOYO (cap+viaje)")
    f_apoyo = formula_cell(ws, row, 3, f"={dias_cap_ref}+{dias_viaje_ref}")
    param_rows["total_apoyo"] = row
    row += 1

    row += 1  # spacer

    # ---- TABLE 2: TARIFAS ----
    section_header(ws, row, 2, "TABLA 2 — TARIFAS TÉCNICOS (S/. por hora)", span=3)
    row += 1
    col_header(ws, row, 2, "Especialidad")
    col_header(ws, row, 3, "Tarifa Lead")
    col_header(ws, row, 4, "Tarifa Apoyo")
    tarifas_header_row = row
    row += 1

    tarifas_start = row
    for i, (esp, tl, ta) in enumerate(TARIFAS):
        bg = C_ALT if i % 2 == 0 else C_WHITE
        for c, v in [(2, esp), (3, tl), (4, ta)]:
            cell = ws.cell(row=row, column=c, value=v)
            cell.fill = fill(bg) if c == 2 else fill(C_YELLOW)
            cell.font = font(size=10)
            cell.alignment = align("right" if c > 2 else "left")
            cell.border = thin_border()
        row += 1
    tarifas_end = row - 1

    row += 1  # spacer

    # ---- TABLE 3: VIÁTICOS ----
    section_header(ws, row, 2, "TABLA 3 — VIÁTICOS Y TRANSPORTE DIARIO (por persona)", span=4)
    row += 1
    col_header(ws, row, 2, "Grupo")
    col_header(ws, row, 3, "Descripción")
    col_header(ws, row, 4, "Viático/día")
    col_header(ws, row, 5, "Transp.interno/día")
    col_header(ws, row, 6, "Noche hospedaje")
    grupos_header_row = row
    row += 1

    grupos_start = row
    for i, (grp, desc, viat, transp, noche) in enumerate(GRUPOS):
        bg = C_ALT if i % 2 == 0 else C_WHITE
        for c, v in [(2, grp), (3, desc), (4, viat), (5, transp), (6, noche)]:
            cell = ws.cell(row=row, column=c, value=v)
            cell.fill = fill(C_YELLOW) if c >= 4 else fill(bg)
            cell.font = font(size=10)
            cell.alignment = align("right" if c >= 4 else "left")
            cell.border = thin_border()
        row += 1
    grupos_end = row - 1

    row += 1

    # ---- TABLE 4: TRASLADO POR IE ----
    section_header(ws, row, 2, "TABLA 4 — TRASLADO POR IE (por persona ida+vuelta)", span=4)
    row += 1
    col_header(ws, row, 2, "IE Código")
    col_header(ws, row, 3, "IE Nombre")
    col_header(ws, row, 4, "Ruta")
    col_header(ws, row, 5, "Medio")
    col_header(ws, row, 6, "Costo S/.")
    row += 1

    colegios_start = row
    for i, (ie_cod, ie_nom, ruta, medio, costo) in enumerate(COLEGIOS):
        bg = C_ALT if i % 2 == 0 else C_WHITE
        for c, v in [(2, ie_cod), (3, ie_nom), (4, ruta), (5, medio), (6, costo)]:
            cell = ws.cell(row=row, column=c, value=v)
            cell.fill = fill(C_YELLOW) if c == 6 else fill(bg)
            cell.font = font(size=10)
            cell.alignment = align("right" if c == 6 else "left")
            cell.border = thin_border()
        row += 1
    colegios_end = row - 1

    row += 1

    # ---- TABLE 5: CONSUMIBLES ----
    section_header(ws, row, 2, "TABLA 5 — CONSUMIBLES POR ESPECIALIDAD (por taller)", span=2)
    row += 1
    col_header(ws, row, 2, "Especialidad")
    col_header(ws, row, 3, "Consumibles S/.")
    row += 1

    consumibles_start = row
    for i, (esp, cons) in enumerate(CONSUMIBLES):
        bg = C_ALT if i % 2 == 0 else C_WHITE
        for c, v in [(2, esp), (3, cons)]:
            cell = ws.cell(row=row, column=c, value=v)
            cell.fill = fill(C_YELLOW) if c == 3 else fill(bg)
            cell.font = font(size=10)
            cell.alignment = align("right" if c == 3 else "left")
            cell.border = thin_border()
        row += 1
    consumibles_end = row - 1

    row += 1

    # ---- TABLE 6: OTROS ----
    section_header(ws, row, 2, "TABLA 6 — OTROS", span=2)
    row += 1
    col_header(ws, row, 2, "Concepto")
    col_header(ws, row, 3, "Monto S/.")
    row += 1

    label_cell(ws, row, 2, "Materiales didácticos por taller")
    editable_cell(ws, row, 3, 150)
    mat_did_row = row
    row += 1

    label_cell(ws, row, 2, "Seguro de viaje por técnico (solo Grupo C)")
    editable_cell(ws, row, 3, 80)
    seguro_row = row
    row += 1

    # Store named ranges as a dict to pass around
    return {
        "ws": ws,
        "param_rows": param_rows,
        "tarifas_start": tarifas_start,
        "tarifas_end": tarifas_end,
        "grupos_start": grupos_start,
        "grupos_end": grupos_end,
        "colegios_start": colegios_start,
        "colegios_end": colegios_end,
        "consumibles_start": consumibles_start,
        "consumibles_end": consumibles_end,
        "mat_did_row": mat_did_row,
        "seguro_row": seguro_row,
    }


# ---------------------------------------------------------------------------
# SHEET 2 — DETALLE M4
# ---------------------------------------------------------------------------
def build_detalle(wb, premisas_info):
    ws = wb.create_sheet("DETALLE M4")
    no_gridlines(ws)

    p = premisas_info
    prem_ws = "PREMISAS"
    pr = p["param_rows"]

    # References to PREMISAS cells
    ref_total_lead  = f"PREMISAS!C{pr['total_lead']}"
    ref_total_apoyo = f"PREMISAS!C{pr['total_apoyo']}"
    ref_dias_cap    = f"PREMISAS!C{pr['dias_cap']}"
    ref_dias_viaje  = f"PREMISAS!C{pr['dias_viaje']}"
    ref_hh_dia      = f"PREMISAS!C{pr['hh_dia']}"
    ref_noches      = f"PREMISAS!C{pr['noches']}"
    ref_mat_did     = f"PREMISAS!C{p['mat_did_row']}"
    ref_seguro      = f"PREMISAS!C{p['seguro_row']}"

    # VLOOKUP ranges in PREMISAS (absolute)
    tar_range  = f"PREMISAS!$B${p['tarifas_start']}:$D${p['tarifas_end']}"
    grp_range  = f"PREMISAS!$B${p['grupos_start']}:$F${p['grupos_end']}"
    col_range  = f"PREMISAS!$B${p['colegios_start']}:$F${p['colegios_end']}"
    cons_range = f"PREMISAS!$B${p['consumibles_start']}:$C${p['consumibles_end']}"

    # Column widths
    col_widths = {
        1: 4,    # A N°
        2: 16,   # B Código
        3: 30,   # C IE Nombre
        4: 7,    # D IE Código
        5: 30,   # E Ubicación
        6: 7,    # F Grupo
        7: 28,   # G Especialidad
        8: 8,    # H Apoyo
        9: 8,    # I Personas
        10: 10,  # J Días Lead
        11: 10,  # K Días Apoyo
        12: 8,   # L HH/día
        13: 12,  # M Tarifa Lead
        14: 12,  # N Tarifa Apoyo
        15: 14,  # O Costo Lead
        16: 14,  # P Costo Apoyo
        17: 15,  # Q Sub Laboral
        18: 12,  # R Viático/día
        19: 10,  # S Días viático
        20: 16,  # T Costo Viáticos
        21: 12,  # U Transp/día
        22: 16,  # V Costo Transp
        23: 16,  # W Traslado/pers
        24: 16,  # X Costo Traslado
        25: 14,  # Y Noche hosp
        26: 8,   # Z Noches
        27: 16,  # AA Costo Hospedaje
        28: 14,  # AB Consumibles
        29: 14,  # AC Mat. didácticos
        30: 18,  # AD TOTAL
    }
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Row 1: Title
    ws.merge_cells("A1:AD1")
    tc = ws["A1"]
    tc.value = "PLAN DE OPERACIONES M4 — DETALLE DE TALLERES"
    tc.fill = fill(C_DARK)
    tc.font = font(bold=True, color=C_WHITE, size=13)
    tc.alignment = align("center")
    ws.row_dimensions[1].height = 28

    # Row 2: Group headers (start_col, span, label)
    # Cols 1-8=IDENTIFICACIÓN, 9-12=TÉCNICOS, 13-17=COSTO LABORAL,
    # 18-20=VIÁTICOS, 21-22=TRANSPORTE INTERNO, 23-24=TRASLADO,
    # 25-27=HOSPEDAJE, 28-29=MATERIALES, 30=TOTAL
    groups = [
        (1,  8,  "IDENTIFICACIÓN"),
        (9,  4,  "TÉCNICOS"),
        (13, 5,  "COSTO LABORAL"),
        (18, 3,  "VIÁTICOS"),
        (21, 2,  "TRANSPORTE INTERNO"),
        (23, 2,  "TRASLADO"),
        (25, 3,  "HOSPEDAJE"),
        (28, 2,  "MATERIALES"),
        (30, 1,  "TOTAL"),
    ]
    for start_col, span, label in groups:
        section_header(ws, 2, start_col, label, span=span)

    # Row 3: Column headers
    headers = [
        "N°", "Código", "IE Nombre", "IE Cód", "Ubicación", "Grupo",
        "Especialidad", "Apoyo\nTéc.", "Personas",
        "Días\nLead", "Días\nApoyo", "HH/día",
        "Tarifa\nLead", "Tarifa\nApoyo",
        "Costo\nLead", "Costo\nApoyo", "SUB\nLABORAL",
        "Viático\n/día", "Días\nViático", "COSTO\nVIÁTICOS",
        "Transp\n/día", "COSTO\nTRANSP",
        "Traslado\n/pers", "COSTO\nTRASLADO",
        "Noche\nHosp", "Noches", "COSTO\nHOSPEDAJE",
        "Consumibles", "Mat.\nDidáct.", "TOTAL",
    ]
    for i, h in enumerate(headers, start=1):
        col_header(ws, 3, i, h)
        ws.row_dimensions[3].height = 36

    # Freeze panes at J4 (freeze rows 1-3 and cols A-I)
    ws.freeze_panes = "J4"

    # Data rows (start at row 4)
    DATA_START = 4
    for t_idx, taller in enumerate(TALLERES):
        row = DATA_START + t_idx
        (n, codigo, ie_nom, ie_cod, ubicacion, grupo, especialidad, apoyo) = taller
        bg = C_ALT if t_idx % 2 == 0 else C_WHITE

        def dc(col, value, fmt=None, bold=False, is_formula=False):
            cell = ws.cell(row=row, column=col, value=value)
            cell.fill = fill(bg)
            cell.font = font(bold=bold, size=9)
            cell.alignment = align("center" if col in [1,6,8,9,10,11,12,19,26] else
                                   "right" if col >= 13 else "left")
            cell.border = thin_border()
            if fmt:
                cell.number_format = fmt
            return cell

        # A-H: static data
        dc(1,  n)
        dc(2,  codigo)
        dc(3,  ie_nom)
        dc(4,  ie_cod)
        dc(5,  ubicacion)
        dc(6,  grupo)
        dc(7,  especialidad)
        dc(8,  apoyo)

        r = row  # shorthand

        # I: Personas
        dc(9,  f'=IF(H{r}="SI",2,1)')

        # J: Días Lead
        dc(10, f"={ref_total_lead}")

        # K: Días Apoyo
        dc(11, f"={ref_total_apoyo}")

        # L: HH/día
        dc(12, f"={ref_hh_dia}")

        # M: Tarifa Lead  (VLOOKUP on especialidad col=G)
        dc(13, f"=VLOOKUP(G{r},{tar_range},2,0)", fmt='S/. #,##0.00')

        # N: Tarifa Apoyo
        dc(14, f"=VLOOKUP(G{r},{tar_range},3,0)", fmt='S/. #,##0.00')

        # O: Costo Lead = J*L*M
        dc(15, f"=J{r}*L{r}*M{r}", fmt='S/. #,##0.00')

        # P: Costo Apoyo = IF(H="SI", K*L*N, 0)
        dc(16, f'=IF(H{r}="SI",K{r}*L{r}*N{r},0)', fmt='S/. #,##0.00')

        # Q: Sub Laboral
        dc(17, f"=O{r}+P{r}", fmt='S/. #,##0.00', bold=True)

        # R: Viático/día = VLOOKUP(F, grupos, col3)
        dc(18, f"=VLOOKUP(F{r},{grp_range},3,0)", fmt='S/. #,##0.00')

        # S: Días viático = dias_cap + dias_viaje
        dc(19, f"={ref_dias_cap}+{ref_dias_viaje}")

        # T: Costo Viáticos = R*S*I
        dc(20, f"=R{r}*S{r}*I{r}", fmt='S/. #,##0.00')

        # U: Transp/día = VLOOKUP(F, grupos, col4)
        dc(21, f"=VLOOKUP(F{r},{grp_range},4,0)", fmt='S/. #,##0.00')

        # V: Costo Transp = U*dias_cap*I
        dc(22, f"=U{r}*{ref_dias_cap}*I{r}", fmt='S/. #,##0.00')

        # W: Traslado/persona = VLOOKUP(D, colegios, col5)
        dc(23, f"=VLOOKUP(D{r},{col_range},5,0)", fmt='S/. #,##0.00')

        # X: Costo Traslado = W*I
        dc(24, f"=W{r}*I{r}", fmt='S/. #,##0.00')

        # Y: Noche hospedaje = VLOOKUP(F, grupos, col5)
        dc(25, f"=VLOOKUP(F{r},{grp_range},5,0)", fmt='S/. #,##0.00')

        # Z: Noches
        dc(26, f"={ref_noches}")

        # AA: Costo Hospedaje = Y*Z*I
        dc(27, f"=Y{r}*Z{r}*I{r}", fmt='S/. #,##0.00')

        # AB: Consumibles = VLOOKUP(G, consumibles, col2)
        dc(28, f"=VLOOKUP(G{r},{cons_range},2,0)", fmt='S/. #,##0.00')

        # AC: Mat. didácticos
        dc(29, f"={ref_mat_did}", fmt='S/. #,##0.00')

        # AD: TOTAL = Q+T+V+X+AA+AB+AC + IF(F="C",I*seguro,0)
        dc(30, f'=Q{r}+T{r}+V{r}+X{r}+AA{r}+AB{r}+AC{r}+IF(F{r}="C",I{r}*{ref_seguro},0)',
           fmt='S/. #,##0.00', bold=True)

    # TOTALS row
    tot_row = DATA_START + len(TALLERES)
    ws.row_dimensions[tot_row].height = 20

    for col in range(1, 31):
        cell = ws.cell(row=tot_row, column=col)
        cell.fill = fill(C_DARK)
        cell.font = font(bold=True, color=C_WHITE, size=10)
        cell.border = thin_border()

    ws.cell(row=tot_row, column=1, value="TOTAL").alignment = align("center")
    sum_cols = [9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30]
    for col in sum_cols:
        cl = get_column_letter(col)
        cell = ws.cell(row=tot_row, column=col,
                       value=f"=SUM({cl}{DATA_START}:{cl}{tot_row-1})")
        cell.fill = fill(C_DARK)
        cell.font = font(bold=True, color=C_WHITE, size=10)
        cell.number_format = 'S/. #,##0.00'
        cell.alignment = align("right")
        cell.border = thin_border()

    return {
        "ws": ws,
        "data_start": DATA_START,
        "tot_row": tot_row,
    }


# ---------------------------------------------------------------------------
# SHEET 3 — RESUMEN IE
# ---------------------------------------------------------------------------
def build_resumen_ie(wb, detalle_info):
    ws = wb.create_sheet("RESUMEN IE")
    no_gridlines(ws)

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 35
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 20

    # Title
    ws.merge_cells("B1:E1")
    tc = ws["B1"]
    tc.value = "RESUMEN DE COSTOS POR IE"
    tc.fill = fill(C_DARK)
    tc.font = font(bold=True, color=C_WHITE, size=13)
    tc.alignment = align("center")
    ws.row_dimensions[1].height = 28

    row = 3
    section_header(ws, row, 2, "RESUMEN POR INSTITUCIÓN EDUCATIVA", span=4)
    row += 1

    for c, h in [(2,"IE Código"),(3,"IE Nombre"),(4,"Grupo"),(5,"TOTAL S/.")]:
        col_header(ws, row, c, h)
    row += 1

    d = detalle_info
    det_sheet = "'DETALLE M4'"
    d_start = d["data_start"]
    d_end   = d["tot_row"] - 1

    # Unique IEs in order
    seen = []
    for t in TALLERES:
        ie = (t[3], t[2])  # (IE cod, IE nom)
        if ie not in seen:
            seen.append(ie)

    ie_grupo = {t[3]: t[5] for t in TALLERES}

    ie_start = row
    for i, (ie_cod, ie_nom) in enumerate(seen):
        bg = C_ALT if i % 2 == 0 else C_WHITE
        grp = ie_grupo[ie_cod]

        for c, v in [(2, ie_cod), (3, ie_nom), (4, grp)]:
            cell = ws.cell(row=row, column=c, value=v)
            cell.fill = fill(bg)
            cell.font = font(size=10)
            cell.alignment = align("left")
            cell.border = thin_border()

        # SUMIF on col D (IE Código = col 4) to sum col AD (col 30)
        cell = ws.cell(row=row, column=5,
                       value=f"=SUMIF('{det_sheet}'!$D${d_start}:$D${d_end},"
                             f"B{row},"
                             f"'{det_sheet}'!$AD${d_start}:$AD${d_end})")
        # Fix double-quoted sheet name
        cell.value = (f"=SUMIF('DETALLE M4'!$D${d_start}:$D${d_end},"
                      f"B{row},"
                      f"'DETALLE M4'!$AD${d_start}:$AD${d_end})")
        cell.fill = fill(bg)
        cell.font = font(size=10)
        cell.alignment = align("right")
        cell.border = thin_border()
        cell.number_format = 'S/. #,##0.00'
        row += 1

    # Total row
    tot_row = row
    for c in range(2, 6):
        cell = ws.cell(row=tot_row, column=c)
        cell.fill = fill(C_DARK)
        cell.font = font(bold=True, color=C_WHITE, size=10)
        cell.border = thin_border()

    ws.cell(row=tot_row, column=2, value="TOTAL").alignment = align("center")
    total_cell = ws.cell(row=tot_row, column=5,
                         value=f"=SUM(E{ie_start}:E{tot_row-1})")
    total_cell.fill = fill(C_DARK)
    total_cell.font = font(bold=True, color=C_WHITE, size=10)
    total_cell.number_format = 'S/. #,##0.00'
    total_cell.alignment = align("right")
    total_cell.border = thin_border()


# ---------------------------------------------------------------------------
# SHEET 4 — RESUMEN ESPECIALIDAD
# ---------------------------------------------------------------------------
def build_resumen_especialidad(wb, detalle_info):
    ws = wb.create_sheet("RESUMEN ESPECIALIDAD")
    no_gridlines(ws)

    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 20

    # Title
    ws.merge_cells("B1:D1")
    tc = ws["B1"]
    tc.value = "RESUMEN DE COSTOS POR ESPECIALIDAD"
    tc.fill = fill(C_DARK)
    tc.font = font(bold=True, color=C_WHITE, size=13)
    tc.alignment = align("center")
    ws.row_dimensions[1].height = 28

    row = 3
    section_header(ws, row, 2, "RESUMEN POR ESPECIALIDAD", span=3)
    row += 1

    for c, h in [(2,"Especialidad"),(3,"# Talleres"),(4,"TOTAL S/.")]:
        col_header(ws, row, c, h)
    row += 1

    d = detalle_info
    d_start = d["data_start"]
    d_end   = d["tot_row"] - 1

    # Unique especialidades sorted
    especialidades = sorted(set(t[6] for t in TALLERES))

    esp_start = row
    for i, esp in enumerate(especialidades):
        bg = C_ALT if i % 2 == 0 else C_WHITE
        count = sum(1 for t in TALLERES if t[6] == esp)

        cell = ws.cell(row=row, column=2, value=esp)
        cell.fill = fill(bg)
        cell.font = font(size=10)
        cell.alignment = align("left")
        cell.border = thin_border()

        cell = ws.cell(row=row, column=3, value=count)
        cell.fill = fill(bg)
        cell.font = font(size=10)
        cell.alignment = align("center")
        cell.border = thin_border()

        # SUMIF on col G (especialidad = col 7) to sum col AD (col 30)
        cell = ws.cell(row=row, column=4,
                       value=(f"=SUMIF('DETALLE M4'!$G${d_start}:$G${d_end},"
                              f"B{row},"
                              f"'DETALLE M4'!$AD${d_start}:$AD${d_end})"))
        cell.fill = fill(bg)
        cell.font = font(size=10)
        cell.alignment = align("right")
        cell.border = thin_border()
        cell.number_format = 'S/. #,##0.00'
        row += 1

    # Total row
    tot_row = row
    for c in range(2, 5):
        cell = ws.cell(row=tot_row, column=c)
        cell.fill = fill(C_DARK)
        cell.font = font(bold=True, color=C_WHITE, size=10)
        cell.border = thin_border()

    ws.cell(row=tot_row, column=2, value="TOTAL").alignment = align("center")

    tot_count = ws.cell(row=tot_row, column=3,
                        value=f"=SUM(C{esp_start}:C{tot_row-1})")
    tot_count.fill = fill(C_DARK)
    tot_count.font = font(bold=True, color=C_WHITE, size=10)
    tot_count.alignment = align("center")
    tot_count.border = thin_border()

    total_cell = ws.cell(row=tot_row, column=4,
                         value=f"=SUM(D{esp_start}:D{tot_row-1})")
    total_cell.fill = fill(C_DARK)
    total_cell.font = font(bold=True, color=C_WHITE, size=10)
    total_cell.number_format = 'S/. #,##0.00'
    total_cell.alignment = align("right")
    total_cell.border = thin_border()


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------
def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Build sheets in order
    premisas_info  = build_premisas(wb)
    detalle_info   = build_detalle(wb, premisas_info)
    build_resumen_ie(wb, detalle_info)
    build_resumen_especialidad(wb, detalle_info)

    out_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                            "Plan_Operaciones_M4.xlsx")
    wb.save(out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
