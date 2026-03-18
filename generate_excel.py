import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Ruta de Aprendizaje"

# ── Colores por módulo ──────────────────────────────────────────────────────
MODULO_COLORS = {
    "M0": "BDE8F0",  # cian suave
    "M1": "FDDADA",  # rojo suave
    "M2": "D0EEF8",  # azul claro
    "M3": "D4F5E2",  # verde claro
    "M4": "E9DCFD",  # lila suave
    "M5": "FEF3C7",  # amarillo suave
    "M6": "C7F9E0",  # verde menta
}

# Colores fila por formato
FORMATO_COLORS = {
    "Video":       "E8F7FF",
    "PDF":         "FFF0F0",
    "Descargable": "F0FFF4",
    "Interactivo": "FFF8E8",
    "Quiz":        "FDF6C0",
    "Zoom / En Vivo": "F0EEFF",
    "Actividad Presencial": "FFFDE0",
}

def fmt_duracion(duracion_min, paginas):
    if duracion_min:
        h = duracion_min // 60
        m = duracion_min % 60
        if h > 0 and m > 0:
            return f"{h}h {m}min"
        elif h > 0:
            return f"{h}h"
        else:
            return f"{m} min"
    elif paginas:
        return f"{paginas} págs."
    return "—"

def fmt_formato(tipo):
    mapping = {
        "VIDEO":              "Video",
        "PDF":                "PDF",
        "DESCARGABLE":        "Descargable",
        "INTERACTIVO":        "Interactivo",
        "QUIZ":               "Quiz",
        "EN_VIVO":            "Zoom / En Vivo",
        "ACTIVIDAD_PRACTICA": "Actividad Presencial",
    }
    return mapping.get(tipo, tipo)

def fmt_modalidad(modalidad):
    mapping = {
        "asincrono":  "Asíncrono",
        "sincrono":   "Sincrónico",
        "presencial": "Presencial",
    }
    return mapping.get(modalidad, modalidad)

# ── Datos completos ─────────────────────────────────────────────────────────
data = [
    # modulo_label, num_sesion, sesion_titulo, sesion_desc, contenido_titulo, contenido_desc, logro, modalidad, duracion_min, paginas, tipo
    # M0
    ("M0 · Inicio y Diagnóstico", "0.1", "Bienvenida y Pacto de Aprendizaje",
     "Presentación del programa, objetivos y compromiso del participante",
     "Bienvenida al Programa TSF — MINEDU 2024",
     "Presentación oficial del programa, los 9 talleres y la hoja de ruta de 150 horas",
     "Conocer la estructura del programa TSF, los 9 talleres y el compromiso de 150 horas",
     "asincrono", 10, None, "VIDEO"),

    ("M0 · Inicio y Diagnóstico", "0.1", "Bienvenida y Pacto de Aprendizaje",
     "Presentación del programa, objetivos y compromiso del participante",
     "Kit del Participante — Descarga e imprime",
     "Guía completa del participante: cronograma, materiales requeridos y pacto de aprendizaje para firmar",
     "Contar con el cronograma, materiales y pacto de aprendizaje firmado",
     "asincrono", None, 12, "DESCARGABLE"),

    ("M0 · Inicio y Diagnóstico", "0.2", "Selección de Perfil de Grado",
     "Indica el grado que enseñarás para personalizar los contenidos",
     "¿Qué grado enseñas?",
     "Selecciona tu grado (1°–5°) para adaptar los ejemplos pedagógicos a tu contexto",
     "Personalizar los contenidos del programa según el grado a cargo",
     "asincrono", 5, None, "INTERACTIVO"),

    ("M0 · Inicio y Diagnóstico", "0.3", "Diagnóstico Situacional Técnico",
     "¿Cuánto sabes ya del equipamiento de tu taller?",
     "Diagnóstico Técnico — Línea de base",
     "8 preguntas situacionales para medir conocimiento previo del equipamiento. Sin nota mínima — solo para calibrar",
     "Identificar el punto de partida técnico del docente respecto al equipamiento del taller",
     "asincrono", 15, None, "QUIZ"),

    ("M0 · Inicio y Diagnóstico", "0.4", "Diagnóstico Pedagógico",
     "¿Cómo planificas hoy? ¿Qué dificultades tienes en el taller?",
     "Diagnóstico Pedagógico — Situaciones de Aula",
     "6 preguntas sobre prácticas pedagógicas actuales. Sirve para calibrar el programa, no para evaluar",
     "Identificar las prácticas pedagógicas actuales y las dificultades en el taller",
     "asincrono", 12, None, "QUIZ"),

    ("M0 · Inicio y Diagnóstico", "0.5", "Tour Virtual del Taller",
     "Recorre el taller completo en 360° antes de conocerlo físicamente",
     "Tour Virtual 360° — Tu Taller Equipado",
     "Recorrido visual de las 4 zonas: Investigación, Innovación, Acabados y Almacén. Con narración técnica de cada equipo",
     "Identificar las 4 zonas del taller y los equipos principales antes de la visita presencial",
     "asincrono", 45, None, "VIDEO"),

    ("M0 · Inicio y Diagnóstico", "0.6", "Sesión Sincrónica de Apertura",
     "Primer encuentro con tu grupo y el formador GRAMA",
     "Sesión de Apertura — Presentación y expectativas",
     "Conoce a tu grupo, comparte expectativas y recibe orientaciones del formador para el desarrollo del programa",
     "Establecer vínculos con el grupo y acordar expectativas del programa",
     "sincrono", 60, None, "EN_VIVO"),

    # M1
    ("M1 · Seguridad y Arquitectura del Taller", "1.1", "Marco Normativo de Seguridad",
     "Ley 29783 y Manual SSO — fundamentos legales",
     "Marco Normativo de Seguridad Ocupacional en Talleres EPT",
     "Ley 29783, Manual SSO MINEDU y obligaciones del docente como responsable del taller",
     "Conocer y aplicar el marco legal de seguridad y sus obligaciones como responsable del taller",
     "asincrono", 45, 28, "PDF"),

    ("M1 · Seguridad y Arquitectura del Taller", "1.2", "EPP por Proceso",
     "Tabla de Equipos de Protección Personal según operación y equipo",
     "Selector de EPP por Equipo y Proceso",
     "Tabla filtrable interactiva: selecciona el equipo o proceso y obtén el EPP exacto requerido",
     "Identificar y seleccionar el EPP correcto para cada operación y equipo del taller",
     "asincrono", 30, None, "INTERACTIVO"),

    ("M1 · Seguridad y Arquitectura del Taller", "1.3", "Protocolos de Seguridad Operativa",
     "Procedimientos paso a paso para el uso seguro de cada equipo",
     "Protocolos de Seguridad — Máquinas y Herramientas",
     "Demostración en video de protocolos: encendido, operación y apagado. Incluye errores comunes y prevención",
     "Aplicar los protocolos de seguridad operativa en el uso de máquinas y herramientas del taller",
     "asincrono", 60, None, "VIDEO"),

    ("M1 · Seguridad y Arquitectura del Taller", "1.3", "Protocolos de Seguridad Operativa",
     "Procedimientos paso a paso para el uso seguro de cada equipo",
     "Fichas de Revisión Diaria — Plastificables A5",
     "Fichas para cada equipo crítico: lista de verificación antes del uso. Para imprimir y plastificar",
     "Implementar un sistema de verificación diaria del estado de las máquinas del taller",
     "asincrono", None, 16, "DESCARGABLE"),

    ("M1 · Seguridad y Arquitectura del Taller", "1.4", "Mapa de Seguridad del Taller",
     "Señalización, rutas de evacuación y ubicación de equipos de emergencia",
     "Constructor de Mapa de Seguridad",
     "Herramienta para crear el mapa de seguridad del taller: señalización, extintores, salidas y zonas de riesgo",
     "Elaborar el mapa de seguridad del taller con señalización normalizada",
     "asincrono", 40, None, "INTERACTIVO"),

    ("M1 · Seguridad y Arquitectura del Taller", "1.5", "Arquitectura y Zonas del Taller",
     "Distribución de las 4 zonas y catálogo completo de bienes",
     "Las 4 Zonas del Taller — Lógica y Diseño",
     "Explicación pedagógica de por qué el taller está organizado en 4 zonas y su vinculación con el currículo EPT",
     "Comprender la lógica pedagógica de las 4 zonas del taller y su relación con el currículo EPT",
     "asincrono", 60, None, "VIDEO"),

    ("M1 · Seguridad y Arquitectura del Taller", "1.5", "Arquitectura y Zonas del Taller",
     "Distribución de las 4 zonas y catálogo completo de bienes",
     "Catálogo Interactivo de Bienes del Taller",
     "Explorador de todos los bienes por zona con descripción, uso pedagógico y normas de seguridad",
     "Identificar todos los bienes del taller, su zona y sus normas de uso pedagógico",
     "asincrono", 30, None, "INTERACTIVO"),

    ("M1 · Seguridad y Arquitectura del Taller", "1.6", "Quiz de Seguridad — OBLIGATORIO (mín. 80%)",
     "Debes obtener 80% para desbloquear M2 y M3",
     "Quiz de Seguridad — 8 preguntas · Mínimo 80%",
     "Evaluación de competencias de seguridad con situaciones reales del taller. Mínimo 80% para continuar",
     "Demostrar competencias de seguridad aplicadas a situaciones reales del taller (mín. 80%)",
     "asincrono", 25, None, "QUIZ"),

    ("M1 · Seguridad y Arquitectura del Taller", "1.7", "Sesión Sincrónica de Cierre M1",
     "Revisión de casos reales y resolución de dudas sobre seguridad",
     "Sesión Cierre M1 — Casos de Seguridad en el Taller",
     "Análisis de incidentes reales en talleres EPT, protocolo ante emergencias y preguntas del grupo",
     "Analizar incidentes reales y consolidar protocolos de respuesta ante emergencias",
     "sincrono", 120, None, "EN_VIVO"),

    ("M1 · Seguridad y Arquitectura del Taller", "1.8", "Visita de Reconocimiento Presencial",
     "Primera visita al taller físico — levantamiento del plano de seguridad",
     "Visita Presencial 1 — Reconocimiento y Plano de Seguridad",
     "Recorre el taller, verifica instalación de equipos, identifica riesgos y elabora el plano de seguridad",
     "Elaborar el plano de seguridad del taller con señalización y rutas de evacuación (entregable: plano A3)",
     "presencial", 120, None, "ACTIVIDAD_PRACTICA"),

    # M2
    ("M2 · Zona de Investigación", "2.1", "Presentación Técnica del Equipamiento",
     "Computadora, impresora 3D, pizarra táctil, tablet, cámara, filmadora, grabadora",
     "Equipos de Investigación — Uso Pedagógico y Técnico",
     "Video por cada equipo de la zona: ficha técnica, normas de operación, uso en proyectos y vinculación con EPT",
     "Operar con seguridad los equipos de la zona de investigación y planificar su uso pedagógico",
     "asincrono", 90, None, "VIDEO"),

    ("M2 · Zona de Investigación", "2.2", "Software como Herramienta Pedagógica",
     "CAD, diseño gráfico, edición de video — herramientas para el aula",
     "Software Creativo en el Taller — CAD, Diseño, Video",
     "Cómo usar el software incluido (Tinkercad, Canva, CapCut) para que estudiantes documenten y diseñen",
     "Incorporar herramientas de software creativo en proyectos estudiantiles del taller",
     "asincrono", 90, None, "VIDEO"),

    ("M2 · Zona de Investigación", "2.3", "Los Lienzos Metodológicos en el Aula",
     "Design Thinking y Lean Canvas adaptados a la EPT",
     "Design Thinking y Lean Canvas — Metodologías para el Taller",
     "Aplicación práctica de metodologías de innovación en proyectos técnicos escolares",
     "Aplicar Design Thinking y Lean Canvas en proyectos técnicos con estudiantes",
     "asincrono", 60, None, "VIDEO"),

    ("M2 · Zona de Investigación", "2.4", "Explorador de Bienes — Zona Investigación",
     "Todos los equipos de la zona con fichas detalladas",
     "Explorador Interactivo — Zona de Investigación",
     "Navega por todos los bienes de la zona vinculado al Repositorio de recursos del taller",
     "Consultar fichas detalladas de todos los bienes de la zona de investigación",
     "asincrono", 25, None, "INTERACTIVO"),

    ("M2 · Zona de Investigación", "2.5", "Quiz Zona Investigación",
     "Evaluación de comprensión del uso pedagógico de los equipos",
     "Quiz — Zona de Investigación · 10 preguntas",
     "Verifica tu comprensión del uso pedagógico de los equipos de investigación. Mínimo 75%",
     "Verificar la comprensión del uso pedagógico de los equipos de la zona de investigación (mín. 75%)",
     "asincrono", 20, None, "QUIZ"),

    ("M2 · Zona de Investigación", "2.6", "Práctica Presencial 1 — Diseño y Prototipado",
     "Sesión en el taller: proyecto de diseño con equipos de investigación",
     "Práctica Presencial S1 — Diseño y Prototipado Digital",
     "Usa cámara, tablet y software de diseño para desarrollar un prototipo digital de proyecto estudiantil",
     "Desarrollar un prototipo digital de proyecto estudiantil usando equipos de investigación (entregable: brief)",
     "presencial", 240, None, "ACTIVIDAD_PRACTICA"),

    ("M2 · Zona de Investigación", "2.7", "Práctica Presencial 2 — Herramientas de Investigación",
     "Aplicación de metodologías de indagación con equipos digitales",
     "Práctica Presencial S2 — Investigación con Herramientas Digitales",
     "Implementa una sesión completa de investigación con estudiantes usando equipos de la zona",
     "Implementar una sesión de investigación documentada con filmadora y cámara (evidencias)",
     "presencial", 240, None, "ACTIVIDAD_PRACTICA"),

    # M3
    ("M3 · Zona de Innovación: Máquinas y Herramientas", "3.A", "Máquinas de Corte y Habilitado",
     "Sierra circular, sierra radial, garlopa, regruesadora",
     "Máquinas de Corte — Operación Segura y Pedagógica",
     "Sierra circular, radial, garlopa y regruesadora: técnica correcta, ajustes y aplicaciones en proyectos",
     "Operar con seguridad las máquinas de corte y habilitado en proyectos estudiantiles",
     "asincrono", 120, None, "VIDEO"),

    ("M3 · Zona de Innovación: Máquinas y Herramientas", "3.B", "Máquinas de Fabricación Digital",
     "Cortadora láser, Router CNC, Impresora 3D",
     "Fabricación Digital — Láser, CNC e Impresión 3D",
     "Flujo completo: desde el diseño en software hasta la pieza terminada. Configuración y parámetros",
     "Ejecutar el flujo completo de fabricación digital: diseño → parámetros → pieza terminada",
     "asincrono", 120, None, "VIDEO"),

    ("M3 · Zona de Innovación: Máquinas y Herramientas", "3.C", "Máquinas de Formado y Armado",
     "Sierra cinta, caladora, torno, taladro de banco, enchapadora",
     "Máquinas de Formado y Armado — Técnicas y Proyectos",
     "Sierra cinta y caladora para cortes curvos, torno para piezas cilíndricas, taladro de banco y enchapadora",
     "Operar máquinas de formado y armado para producir piezas en proyectos estudiantiles",
     "asincrono", 90, None, "VIDEO"),

    ("M3 · Zona de Innovación: Máquinas y Herramientas", "3.D", "Herramientas Manuales",
     "Formones, escofinas, gubias — tallado y acabado manual",
     "Herramientas Manuales — Dominio Técnico y Pedagógico",
     "Uso correcto de formones, escofinas, gubias. Afilado, mantenimiento y aplicaciones en acabado",
     "Usar correctamente herramientas manuales de filo con seguridad en proyectos de acabado",
     "asincrono", 60, None, "VIDEO"),

    ("M3 · Zona de Innovación: Máquinas y Herramientas", "3.E", "Quiz Zona Innovación — EVALUACIÓN (mín. 80%)",
     "Preguntas de seguridad específicas por máquina. Mínimo 80%",
     "Quiz Zona Innovación — 20 preguntas · Mínimo 80%",
     "Evaluación de conocimiento técnico y de seguridad de máquinas con situaciones reales",
     "Demostrar dominio técnico y de seguridad de las máquinas de la zona de innovación (mín. 80%)",
     "asincrono", 35, None, "QUIZ"),

    ("M3 · Zona de Innovación: Máquinas y Herramientas", "3.F", "Práctica Presencial — Nivel 1 Corte y Habilitado",
     "2 sesiones presenciales con sierra circular, radial y garlopa",
     "Práctica Nivel 1 — Corte y Habilitado (2 sesiones · 8h)",
     "Domina las máquinas de corte aplicando protocolos de seguridad. Produce piezas con precisión",
     "Producir piezas de madera habilitadas con precisión usando máquinas de corte (entregable: set de piezas + ficha de proceso)",
     "presencial", 480, None, "ACTIVIDAD_PRACTICA"),

    ("M3 · Zona de Innovación: Máquinas y Herramientas", "3.G", "Práctica Presencial — Nivel 1 Digital y Formado",
     "2 sesiones con CNC, cortadora láser e impresora 3D",
     "Práctica Nivel 1 — Fabricación Digital y Formado (2 sesiones · 8h)",
     "Produce una pieza con cortadora láser o router CNC y una impresa en 3D. Documenta resultados",
     "Producir piezas con fabricación digital y documentar el proceso (entregable: pieza CNC/láser + impresa 3D)",
     "presencial", 480, None, "ACTIVIDAD_PRACTICA"),

    ("M3 · Zona de Innovación: Máquinas y Herramientas", "3.H", "Práctica Presencial — Nivel 2 Ensamble",
     "Integración de procesos: de la pieza al producto terminado",
     "Práctica Nivel 2 — Ensamble Integrado (1 sesión · 6h)",
     "Integra piezas cortadas, maquinadas y formadas en un producto final. Aplica técnicas de ensamble",
     "Integrar piezas de distintos procesos en un producto ensamblado con documentación técnica",
     "presencial", 360, None, "ACTIVIDAD_PRACTICA"),

    # M4
    ("M4 · Acabados y Almacén", "4.1", "Equipamiento de Acabados",
     "Lijadoras, compresor, pistola de pintar, cabina de pintura",
     "Zona de Acabados — Equipos y Técnicas de Aplicación",
     "Lijadoras, compresor, pistola de pintar y cabina: uso correcto, mezcla de pinturas, EPP y mantenimiento",
     "Operar los equipos de acabados con seguridad y aplicar técnicas profesionales de acabado",
     "asincrono", 75, None, "VIDEO"),

    ("M4 · Acabados y Almacén", "4.2", "Fichas de Revisión Diaria",
     "Instrumentos de control preventivo por equipo crítico",
     "Pack de Fichas de Revisión Diaria — A4 Plastificable",
     "Una ficha por equipo crítico para imprimir en A4, plastificar y colgar en cada máquina",
     "Implementar un sistema de control preventivo diario en todos los equipos críticos del taller",
     "asincrono", None, 24, "DESCARGABLE"),

    ("M4 · Acabados y Almacén", "4.3", "Gestión del Almacén",
     "Organización, control de stock y registro de herramientas",
     "Gestión del Almacén del Taller — Sistema de Control",
     "Organización por categorías, sistema de préstamo y devolución, control de stock y registro de inventario",
     "Implementar un sistema organizado de gestión del almacén con control de stock y préstamo",
     "asincrono", 90, None, "VIDEO"),

    ("M4 · Acabados y Almacén", "4.3", "Gestión del Almacén",
     "Organización, control de stock y registro de herramientas",
     "Manual de Organización del Almacén EPT",
     "Guía completa: distribución física, etiquetado, registro digital y reportes de estado para la UGEL",
     "Organizar el almacén según los estándares UGEL con distribución, etiquetado y registro digital",
     "asincrono", None, 20, "PDF"),

    ("M4 · Acabados y Almacén", "4.4", "Mantenimiento Preventivo",
     "Plan de mantenimiento mensual, trimestral y anual del taller",
     "Mantenimiento Preventivo — Antes de que Falle",
     "Plan de mantenimiento por tipo de equipo: lubricación, calibración, limpieza y revisión de seguridad",
     "Diseñar e implementar un plan de mantenimiento preventivo del taller",
     "asincrono", 90, None, "VIDEO"),

    ("M4 · Acabados y Almacén", "4.4", "Mantenimiento Preventivo",
     "Plan de mantenimiento mensual, trimestral y anual del taller",
     "Bitácora del Taller — Registro de Mantenimiento",
     "Plantilla de bitácora anual para registrar mantenimientos, incidentes, préstamos y estado del equipamiento",
     "Llevar una bitácora de mantenimiento del taller en formato oficial aceptado por UGEL",
     "asincrono", None, 8, "DESCARGABLE"),

    ("M4 · Acabados y Almacén", "4.5", "Quiz Acabados y Almacén",
     "Evaluación del sistema de acabados, gestión del almacén y mantenimiento",
     "Quiz — Acabados y Almacén · 12 preguntas",
     "Verifica comprensión del sistema de acabados, gestión del almacén y mantenimiento. Mínimo 75%",
     "Verificar el dominio de acabados, gestión del almacén y mantenimiento preventivo (mín. 75%)",
     "asincrono", 20, None, "QUIZ"),

    ("M4 · Acabados y Almacén", "4.6", "Práctica Presencial — Acabado Completo",
     "Aplica acabados profesionales a las piezas ensambladas en M3",
     "Práctica Presencial — Acabado Completo (4h)",
     "Aplica lijado, sellado y pintura/barniz al producto del M3. Usa la cabina de pintura con EPP completo",
     "Aplicar acabados profesionales a un producto del taller (entregable: producto terminado documentado)",
     "presencial", 240, None, "ACTIVIDAD_PRACTICA"),

    ("M4 · Acabados y Almacén", "4.7", "Práctica Presencial — Gestión del Taller",
     "Implementa el sistema de almacén y fichas de control en tu taller",
     "Práctica Presencial — Sistema de Gestión del Taller (3h)",
     "Organiza el almacén, coloca fichas en máquinas, inicia la bitácora y presenta el sistema al supervisor",
     "Implementar el sistema de gestión del taller (entregable: taller con gestión activa)",
     "presencial", 180, None, "ACTIVIDAD_PRACTICA"),

    # M5
    ("M5 · Programa Formativo en el Taller", "5.1", "El Taller como Espacio de Competencias",
     "Cómo el nuevo equipamiento transforma la enseñanza EPT",
     "El Taller Equipado como Espacio de Aprendizaje",
     "Caso práctico real: docente con el nuevo equipamiento, su primera sesión y los cambios en el aprendizaje",
     "Comprender cómo el equipamiento transforma el espacio de aprendizaje técnico",
     "asincrono", 60, None, "VIDEO"),

    ("M5 · Programa Formativo en el Taller", "5.2", "Las 14 Habilidades desde el Taller",
     "Mapa completo de habilidades EPT vinculadas al equipamiento",
     "Mapa de 14 Habilidades EPT × Equipamiento",
     "Selecciona una habilidad y ve qué equipos la desarrollan. Selecciona un equipo y ve qué habilidades activa",
     "Vincular las 14 habilidades EPT con los equipos específicos del taller",
     "asincrono", 30, None, "INTERACTIVO"),

    ("M5 · Programa Formativo en el Taller", "5.3", "El Itinerario y la Progresión del Equipamiento",
     "Cómo los equipos se secuencian a lo largo de 5 grados",
     "Itinerario Formativo — Cómo Progresa el Uso del Equipamiento",
     "La lógica de progresión por grado: qué equipos se usan en 1°, qué se agrega en 2°, hasta 5°",
     "Comprender la progresión del uso del equipamiento a lo largo de los 5 grados de secundaria",
     "asincrono", 90, None, "VIDEO"),

    ("M5 · Programa Formativo en el Taller", "5.3", "El Itinerario y la Progresión del Equipamiento",
     "Cómo los equipos se secuencian a lo largo de 5 grados",
     "Tabla de Progresión por Grado — Interactiva",
     "Visualiza qué equipos y habilidades corresponden a cada grado y cómo se encadenan los proyectos",
     "Consultar la progresión de equipos y habilidades por grado para la planificación anual",
     "asincrono", 20, None, "INTERACTIVO"),

    ("M5 · Programa Formativo en el Taller", "5.4", "Planificar con el Equipamiento como Ancla",
     "Método de planificación desde el equipo hacia el currículo",
     "Guía de Planificación con Equipamiento — Método GRAMA",
     "Metodología paso a paso: identificar el bien ancla, seleccionar competencias, diseñar actividades",
     "Diseñar unidades de aprendizaje usando el equipamiento como ancla curricular (Método GRAMA)",
     "asincrono", 120, 32, "PDF"),

    ("M5 · Programa Formativo en el Taller", "5.4", "Planificar con el Equipamiento como Ancla",
     "Método de planificación desde el equipo hacia el currículo",
     "Plantilla de Unidad de Aprendizaje EPT — Descargable",
     "Formato oficial de unidad de aprendizaje adaptado al taller equipado. Con ejemplo completo de llenado",
     "Contar con la plantilla oficial para diseñar unidades de aprendizaje EPT",
     "asincrono", None, 4, "DESCARGABLE"),

    ("M5 · Programa Formativo en el Taller", "5.5", "Evaluar Competencias en el Taller",
     "Instrumentos de evaluación técnica y procedimental",
     "Evaluación de Competencias Técnicas — Rúbricas y Listas",
     "Cómo evaluar el desempeño práctico: rúbricas por equipo, listas de cotejo y portafolio digital",
     "Diseñar y aplicar instrumentos de evaluación de competencias técnicas en el taller",
     "asincrono", 90, None, "VIDEO"),

    ("M5 · Programa Formativo en el Taller", "5.5", "Evaluar Competencias en el Taller",
     "Instrumentos de evaluación técnica y procedimental",
     "Pack de Instrumentos de Evaluación — EPT Talleres",
     "Rúbricas, listas de cotejo y fichas de observación para evaluar competencias técnicas",
     "Contar con instrumentos de evaluación adaptables a cualquier taller y grado",
     "asincrono", None, 18, "DESCARGABLE"),

    ("M5 · Programa Formativo en el Taller", "5.6", "El Módulo de Emprendimiento Integrado",
     "Cómo articular el taller con el módulo de emprendimiento",
     "Emprendimiento desde el Taller — Proyectos con Valor",
     "Integración del módulo de emprendimiento con proyectos del taller: de aprender a hacer a crear valor",
     "Articular el módulo de emprendimiento con proyectos técnicos que generen valor de mercado",
     "asincrono", 90, None, "VIDEO"),

    ("M5 · Programa Formativo en el Taller", "5.7", "Sesión Sincrónica 1 — Planificación Colaborativa",
     "Diseño colaborativo de unidades de aprendizaje con el formador",
     "Sesión 1 — Planificación Colaborativa con el Equipamiento",
     "En grupos, diseña una unidad de aprendizaje usando el equipamiento. El formador retroalimenta en tiempo real",
     "Diseñar colaborativamente una unidad de aprendizaje con retroalimentación del formador",
     "sincrono", 120, None, "EN_VIVO"),

    ("M5 · Programa Formativo en el Taller", "5.8", "Sesión Sincrónica 2 — Programa con Equipamiento Real",
     "Ajuste de la planificación al equipamiento real disponible",
     "Sesión 2 — Programa con el Equipamiento Real del Taller",
     "Ajusta la planificación a los equipos reales disponibles. Intercambio con docentes de otras instituciones",
     "Adaptar la planificación al equipamiento real del taller y compartir experiencias entre docentes",
     "sincrono", 120, None, "EN_VIVO"),

    ("M5 · Programa Formativo en el Taller", "5.9", "Sesión Sincrónica 3 — Casos Difíciles",
     "Estrategias de contingencia ante situaciones complejas",
     "Sesión 3 — Casos Difíciles de Planificación",
     "¿Qué haces cuando el equipo falla? Análisis de situaciones complejas y estrategias de contingencia",
     "Desarrollar estrategias de contingencia pedagógica ante fallas del equipamiento",
     "sincrono", 120, None, "EN_VIVO"),

    ("M5 · Programa Formativo en el Taller", "5.10", "Producción de Planificación del Grado",
     "Entregable: unidad de aprendizaje completa para tu grado",
     "Entregable M5 — Planificación Completa del Grado",
     "Produce una unidad de aprendizaje completa para tu grado: situación significativa, actividades, evaluación",
     "Producir una unidad de aprendizaje completa con equipamiento como ancla (entregable M5)",
     "asincrono", 180, None, "ACTIVIDAD_PRACTICA"),

    # M6
    ("M6 · Proyecto Integrador", "6.1", "Briefing del Proyecto Integrador",
     "Rúbrica, criterios de evaluación y selección del proyecto",
     "Briefing del Proyecto Integrador — Rúbrica y Criterios",
     "Especificaciones: qué debe producir, cómo se evaluará, qué procesos integrar y qué documentar",
     "Comprender los criterios y especificaciones del proyecto integrador para planificar su ejecución",
     "asincrono", 60, 8, "PDF"),

    ("M6 · Proyecto Integrador", "6.1", "Briefing del Proyecto Integrador",
     "Rúbrica, criterios de evaluación y selección del proyecto",
     "Rúbrica de Evaluación — Proyecto Integrador",
     "Rúbrica oficial de evaluación con criterios técnicos, pedagógicos y de proceso. En Excel y PDF",
     "Conocer los criterios de evaluación del proyecto integrador y autoevaluar el avance",
     "asincrono", None, 2, "DESCARGABLE"),

    ("M6 · Proyecto Integrador", "6.2", "Sesión Sincrónica — Briefing Colectivo",
     "Presentación y aprobación de propuestas de proyecto",
     "Sesión Briefing — Presentación de Proyectos",
     "Cada participante presenta su propuesta. El grupo y el formador retroalimentan y aprueban el proyecto",
     "Presentar y validar la propuesta de proyecto integrador con retroalimentación del grupo y formador",
     "sincrono", 60, None, "EN_VIVO"),

    ("M6 · Proyecto Integrador", "6.3", "Sesión Presencial 1 — Diseño y Prototipado",
     "Diseño en software y plan de producción del proyecto",
     "Sesión 1 del Proyecto — Diseño y Prototipado (4h)",
     "Diseña el producto final en CAD/gráfico, produce un prototipo digital y elabora el plan de producción",
     "Producir el diseño y prototipo digital del proyecto con su plan de producción (entregable: diseño aprobado)",
     "presencial", 240, None, "ACTIVIDAD_PRACTICA"),

    ("M6 · Proyecto Integrador", "6.4", "Sesión Presencial 2 — Habilitado y Maquinado",
     "Cortes, perforaciones y operaciones de maquinado del proyecto",
     "Sesión 2 del Proyecto — Habilitado y Maquinado (4h)",
     "Ejecuta los cortes, perforaciones y operaciones de maquinado del proyecto con protocolos de seguridad",
     "Ejecutar las operaciones de habilitado y maquinado del proyecto con seguridad",
     "presencial", 240, None, "ACTIVIDAD_PRACTICA"),

    ("M6 · Proyecto Integrador", "6.5", "Sesión Presencial 3 — Ensamble",
     "Ensamble de todas las piezas del proyecto",
     "Sesión 3 del Proyecto — Ensamble (4h)",
     "Ensambla todas las piezas del proyecto. Verifica dimensiones, ajustes y funcionamiento",
     "Ensamblar el proyecto verificando dimensiones y funcionamiento con documentación fotográfica",
     "presencial", 240, None, "ACTIVIDAD_PRACTICA"),

    ("M6 · Proyecto Integrador", "6.6", "Sesión Presencial 4 — Acabado y Documentación",
     "Acabados finales y inicio del expediente pedagógico",
     "Sesión 4 del Proyecto — Acabado Final y Documentación (4h)",
     "Aplica acabados finales al proyecto. Fotografía el producto terminado. Inicia el expediente pedagógico",
     "Aplicar acabados finales y documentar el producto terminado para el expediente pedagógico",
     "presencial", 240, None, "ACTIVIDAD_PRACTICA"),

    ("M6 · Proyecto Integrador", "6.7", "Elaboración del Expediente Pedagógico",
     "Documento técnico-pedagógico del proyecto completo",
     "Expediente Pedagógico — Documentación del Proyecto (4h)",
     "Elabora el expediente: ficha técnica, proceso documentado, instrumentos, evidencias fotográficas y propuesta pedagógica",
     "Elaborar el expediente pedagógico completo del proyecto integrador (entregable final)",
     "presencial", 240, None, "ACTIVIDAD_PRACTICA"),

    ("M6 · Proyecto Integrador", "6.8", "Sustentación y Ceremonia de Cierre",
     "Presentación final, evaluación y certificación del programa",
     "Sustentación y Ceremonia de Certificación",
     "Presenta tu proyecto, defiende decisiones técnicas y pedagógicas, recibe retroalimentación y certifícate",
     "Sustentar el proyecto integrador, recibir evaluación del jurado y obtener la certificación del programa",
     "sincrono", 120, None, "EN_VIVO"),
]

# ── Encabezados ─────────────────────────────────────────────────────────────
HEADERS = [
    "Módulo",
    "N° Sesión",
    "Sesión / Sub-sección",
    "Descripción de la Sesión",
    "Título del Contenido",
    "Descripción del Contenido",
    "Logro de Aprendizaje",
    "Modalidad",
    "Duración",
    "Formato",
]

WIDTHS = [30, 10, 38, 50, 52, 65, 65, 15, 15, 22]

# ── Estilos ──────────────────────────────────────────────────────────────────
header_fill  = PatternFill("solid", fgColor="043941")
header_font  = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
cell_align   = Alignment(vertical="top", wrap_text=True)
thin         = Side(style="thin", color="DDDDDD")
border       = Border(left=thin, right=thin, top=thin, bottom=thin)

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def write_header(ws):
    for col, (h, w) in enumerate(zip(HEADERS, WIDTHS), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

def write_row(ws, row_num, record):
    (modulo_label, num, ses_titulo, ses_desc, cont_titulo, cont_desc,
     logro, modalidad, dur_min, paginas, tipo) = record

    m_code = modulo_label.split("·")[0].strip()  # "M0", "M1", ...
    formato = fmt_formato(tipo)
    duracion = fmt_duracion(dur_min, paginas)
    modal    = fmt_modalidad(modalidad)

    row_color = FORMATO_COLORS.get(formato, "FFFFFF")

    values = [
        modulo_label, num, ses_titulo, ses_desc,
        cont_titulo, cont_desc, logro, modal, duracion, formato
    ]

    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.fill      = make_fill(row_color)
        cell.alignment = cell_align
        cell.border    = border
        cell.font      = Font(name="Calibri", size=9)

    # columna módulo en negrita con color del módulo
    mod_cell = ws.cell(row=row_num, column=1)
    mod_cell.fill = make_fill(MODULO_COLORS.get(m_code, "FFFFFF"))
    mod_cell.font = Font(bold=True, name="Calibri", size=9)

    # num sesion centrado
    ws.cell(row=row_num, column=2).alignment = Alignment(horizontal="center", vertical="top")
    # modalidad centrada
    ws.cell(row=row_num, column=8).alignment = Alignment(horizontal="center", vertical="top")
    # duración centrada
    ws.cell(row=row_num, column=9).alignment = Alignment(horizontal="center", vertical="top")
    # formato centrado y negrita
    fmt_cell = ws.cell(row=row_num, column=10)
    fmt_cell.font = Font(bold=True, name="Calibri", size=9)
    fmt_cell.alignment = Alignment(horizontal="center", vertical="top")

    ws.row_dimensions[row_num].height = 50

write_header(ws)
for i, record in enumerate(data, start=2):
    write_row(ws, i, record)

# ── Segunda hoja: Resumen por módulo ────────────────────────────────────────
ws2 = wb.create_sheet("Resumen por Módulo")
resumen_headers = ["Módulo", "Fase", "Total Horas", "Asíncrono", "Sincrónico", "Presencial", "N° Contenidos"]
modulos_info = [
    ("M0 · Inicio y Diagnóstico",                   "Diagnóstico",    4,  3, 1, 0, 8),
    ("M1 · Seguridad y Arquitectura del Taller",     "Orientación",   13,  9, 2, 2, 11),
    ("M2 · Zona de Investigación",                   "Apropiación",   18, 10, 0, 8, 8),
    ("M3 · Zona de Innovación: Máquinas y Herrms.",  "Aplicación",    36, 14, 0, 22, 9),
    ("M4 · Acabados y Almacén",                      "Aplicación",    14,  7, 0, 7, 9),
    ("M5 · Programa Formativo en el Taller",         "Aplicación",    22, 16, 6, 0, 13),
    ("M6 · Proyecto Integrador",                     "Proyecto",      25,  2, 3, 20, 9),
    ("TOTAL",                                        "",             132, 61, 12, 59, 67),
]

for col, h in enumerate(resumen_headers, 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = header_align
    cell.border = border

ws2.column_dimensions["A"].width = 42
for col in ["B","C","D","E","F","G"]:
    ws2.column_dimensions[col].width = 16

m_codes = ["M0","M1","M2","M3","M4","M5","M6",""]
for row_num, (info, mcode) in enumerate(zip(modulos_info, m_codes + [""]), 2):
    bg = MODULO_COLORS.get(mcode, "F5F5F5") if mcode else "C7F9E0"
    for col, val in enumerate(info, 1):
        cell = ws2.cell(row=row_num, column=col, value=val)
        cell.fill = make_fill(bg)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(name="Calibri", size=10,
                         bold=(col == 1 or info[0] == "TOTAL"))
    ws2.row_dimensions[row_num].height = 22

# Guardar
out = "/home/user/CLAUDE-CODE-GRAMA/Ruta_de_Aprendizaje_GRAMA.xlsx"
wb.save(out)
print(f"✓ Excel generado: {out}")
print(f"  Filas de contenido: {len(data)}")
